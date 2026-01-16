import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# FR Data
fr_data = [
    ['FR101', 'File Upload / Import', 'The application will need to have the ability for users to upload CSV data files.', 'The objective can support the uploading of CSV data files', 'Users can select a CSV data file to upload to the application', 'H', 'Completed'],
    ['FR102', 'Drag and Drop', 'The application will allow users to drag and drop CSV data files to upload.', 'Support drag and drop upload', 'Users can drag and drop a CSV data file from their desktop', 'L', 'Cancelled'],
    ['FR201', 'Data Parsing', 'Parse data files content identifying headers, tables, text', 'Accurately parse data and identify structures', 'Headers, tables and text are correctly identified', 'H', 'Completed'],
    ['FR202', 'Data Extraction', 'Extract data from the CSV file', 'Automatically extract data based on predefined rules', 'Data is successfully extracted from CSV file', 'H', 'Completed'],
    ['FR203', 'Data Transformation', 'Clean and transform data for consistency', 'Format and clean data ready for display', 'Cleaned data has unwanted characters removed', 'H', 'Completed'],
    ['FR301', 'Data Format Selection', 'Map IVD data to predefined template based on user input', 'Allow users to select lab options', 'Users select option from dropdown list', 'M', 'Completed'],
    ['FR302', 'Data Format Determination', 'Read format of uploaded data and determine mapping', 'Auto determine location of IVD test data', 'Data format accurately determined', 'L', 'Completed'],
    ['FR303', 'Data Mapping', 'Map IVDtest data to predefined template', 'Allow users to define data mapping', 'Data categorised based on user defined layout', 'H', 'Completed'],
    ['FR401', 'Data Display and User Interface', 'Show test results to the user', 'Display data in clean, readable tabular format', 'Data displayed in clean format', 'H', 'Completed'],
    ['FR402', 'Result Indication', 'Identify results out of expected range', 'Highlight data points outside expected results', 'Data highlighted when below/above expected', 'L', 'Cancelled'],
    ['FR403', 'Target Selection', 'Allow users to pick targets from filter', 'Allow users to select specific targets', 'Users select targets from dropdown', 'H', 'Completed'],
]

# NFR Data
nfr_data = [
    ['NFR701', 'Ease of Use', 'User-friendly with intuitive understanding of functions', 'M', 'Completed'],
    ['NFR702', 'Learnability', 'Clear and concise documentation for ease of use', 'L', 'Completed'],
    ['NFR801', 'Maintainability', 'Easy maintenance with modular design and documentation', 'M', 'Completed'],
    ['NFR802', 'Data Integrity', 'Sample Name and CT data should not be altered', 'H', 'Completed'],
    ['NFR803', 'Security', 'Should not be hosted online and inherit existing risks', 'H', 'Completed'],
]

# Create FR sheet
fr_sheet = wb.create_sheet('Requirement', 0)
fr_headers = ['Requirement ID', 'Title', 'Description', 'Objective', 'Acceptance Criteria', 'Priority', 'Status']

for col, header in enumerate(fr_headers, 1):
    cell = fr_sheet.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row_idx, row_data in enumerate(fr_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = fr_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Set column widths
fr_sheet.column_dimensions['A'].width = 12
fr_sheet.column_dimensions['B'].width = 20
fr_sheet.column_dimensions['C'].width = 30
fr_sheet.column_dimensions['D'].width = 25
fr_sheet.column_dimensions['E'].width = 25
fr_sheet.column_dimensions['F'].width = 10
fr_sheet.column_dimensions['G'].width = 12

# Create NFR sheet
nfr_sheet = wb.create_sheet('Non-Requirement', 1)
nfr_headers = ['Requirement ID', 'Title', 'Description', 'Priority', 'Status']

for col, header in enumerate(nfr_headers, 1):
    cell = nfr_sheet.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row_idx, row_data in enumerate(nfr_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = nfr_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Set column widths
nfr_sheet.column_dimensions['A'].width = 12
nfr_sheet.column_dimensions['B'].width = 20
nfr_sheet.column_dimensions['C'].width = 45
nfr_sheet.column_dimensions['D'].width = 10
nfr_sheet.column_dimensions['E'].width = 12

wb.save('requirements.xlsx')
print('Excel file created successfully: requirements.xlsx')
