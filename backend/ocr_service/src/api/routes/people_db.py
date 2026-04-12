"""
People Database API Routes - FastAPI endpoints for import and search
"""
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from pydantic import BaseModel
import io
import openpyxl
import csv
import fitz
import re
from datetime import datetime

try:
    from core.people_db_client import get_people_db_client, resolve_quality_thresholds
    from core.supabase_client import get_supabase_client
except ModuleNotFoundError:
    from src.core.people_db_client import get_people_db_client, resolve_quality_thresholds
    from src.core.supabase_client import get_supabase_client
from loguru import logger

router = APIRouter(prefix="/api/v1/people-db", tags=["people-db"])
MAX_IMPORT_FILE_SIZE_BYTES = 25 * 1024 * 1024


# ============================================================================
# Pydantic Models
# ============================================================================

class ImportPreviewResponse(BaseModel):
    """Response for import/preview endpoint compatible with superadmin UI."""
    columns: List[dict]
    row_count: int
    preview_rows: List[dict]


class FieldMapping(BaseModel):
    """Field mapping configuration"""
    nameColumn: int | str
    idNumberColumn: Optional[int | str] = None
    phoneColumn: Optional[int | str] = None
    addressColumn: Optional[int | str] = None
    organizationColumn: Optional[int | str] = None
    positionColumn: Optional[int | str] = None


class ImportSubmitRequest(BaseModel):
    """Request for submitting import (supports both legacy and UI payloads)."""
    importBatchLabel: Optional[str] = None
    dataSource: Optional[str] = None
    fieldMapping: Optional[FieldMapping] = None
    decisions: List[dict] = []

    file_name: Optional[str] = None
    total_rows: Optional[int] = None
    column_mapping: Optional[dict] = None
    data_source: Optional[str] = None
    batch_label: Optional[str] = None


class PeopleSearchResult(BaseModel):
    """Search result for a person"""
    record_id: str
    full_name: str
    id_number: Optional[str] = None
    phone: Optional[str] = None
    mobile: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    company: Optional[str] = None
    data_source: str
    quality_score: float
    ocr_confidence: float
    similarity: float
    import_batch_id: Optional[str] = None
    source_file_path: Optional[str] = None
    source_document_id: Optional[str] = None
    created_at: Optional[str] = None


class PeopleSearchResponse(BaseModel):
    """Response for search endpoint"""
    total: int
    results: List[PeopleSearchResult]
    page: int
    page_size: int
    took_ms: int


class DataSourceFacet(BaseModel):
    key: str
    count: int


class DataSourceFacetsResponse(BaseModel):
    datasets: List[DataSourceFacet]


class ImportBatchSummary(BaseModel):
    batch_id: str
    label: Optional[str] = None
    data_source: Optional[str] = None
    status: Optional[str] = None
    total_records: int = 0
    processed_records: int = 0
    skipped_records: int = 0
    imported_by: Optional[str] = None
    created_at: Optional[str] = None
    updated_at: Optional[str] = None
    error_message: Optional[str] = None


# ============================================================================
# Helper Functions
# ============================================================================

def parse_column_reference(ref: int | str) -> int:
    """Convert column reference to index. 'A' -> 0, 'B' -> 1, etc."""
    if isinstance(ref, int):
        return ref
    # Excel column letter to index
    col_str = str(ref).upper()
    result = 0
    for char in col_str:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1


def bytes_to_mb_text(size_bytes: int) -> str:
    """Convert bytes to readable MB text."""
    return f"{size_bytes / 1024 / 1024:.1f}MB"


def extract_excel_preview(file_content: bytes, max_preview_rows: int = 5) -> tuple:
    """Extract preview from Excel file"""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        worksheet = workbook.active
        
        rows = []
        empty_rows = 0
        
        for idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if idx > max_preview_rows + 10:  # Read a bit more to count empty rows
                break
            
            if not any(cell for cell in row):
                empty_rows += 1
                continue
            
            rows.append(row)
        
        total_rows = worksheet.max_row
        return rows, total_rows, empty_rows
    except Exception as e:
        logger.error(f"Failed to parse Excel: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")


def extract_csv_preview(file_content: bytes, max_preview_rows: int = 5) -> tuple:
    """Extract preview from CSV file"""
    try:
        text_content = file_content.decode('utf-8')
        reader = csv.reader(text_content.split('\n'))
        rows = []
        empty_rows = 0
        
        for idx, row in enumerate(reader):
            if idx > max_preview_rows + 10:
                break
            
            if not any(cell.strip() for cell in row):
                empty_rows += 1
                continue
            
            rows.append(row)
        
        total_rows = len(text_content.split('\n'))
        return rows, total_rows, empty_rows
    except Exception as e:
        logger.error(f"Failed to parse CSV: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse CSV file: {str(e)}")


def split_pdf_text_line(line: str) -> list[str]:
    """Split a PDF text line into columns if common delimiters are present."""
    if '\t' in line:
        cells = [cell.strip() for cell in line.split('\t')]
    elif ',' in line:
        cells = [cell.strip() for cell in line.split(',')]
    elif '|' in line:
        cells = [cell.strip() for cell in line.split('|')]
    elif ';' in line:
        cells = [cell.strip() for cell in line.split(';')]
    else:
        # Some PDF exports separate columns using multiple spaces.
        cells = [cell.strip() for cell in re.split(r'\s{2,}', line)]

    normalized_cells = [cell for cell in cells if cell]
    return normalized_cells if normalized_cells else [line.strip()]


def extract_pdf_preview(file_content: bytes, max_preview_rows: int = 5) -> tuple:
    """Extract text preview rows from PDF file."""
    try:
        doc = fitz.open(stream=file_content, filetype="pdf")
        rows: list[list[str]] = []
        total_rows = 0
        empty_rows = 0
        max_scan_rows = max_preview_rows + 10

        for page in doc:
            page_text = page.get_text("text")
            for raw_line in page_text.splitlines():
                line = raw_line.strip()
                if not line:
                    empty_rows += 1
                    continue

                total_rows += 1
                if len(rows) < max_scan_rows:
                    rows.append(split_pdf_text_line(line))

            if len(rows) >= max_scan_rows:
                continue

        doc.close()
        return rows, total_rows, empty_rows
    except Exception as e:
        logger.error(f"Failed to parse PDF: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse PDF file: {str(e)}")


def build_table_preview(rows: list, has_header: bool, max_preview_rows: int = 5) -> tuple[list[dict], list[dict], int]:
    """Build columns + preview rows from parsed table rows."""
    normalized_rows: list[list] = []
    for row in rows:
        if isinstance(row, (list, tuple)):
            normalized_rows.append(list(row))
        else:
            normalized_rows.append([row])

    if not normalized_rows:
        return [], [], 0

    data_rows = normalized_rows

    if has_header and normalized_rows:
        header_source = normalized_rows[0]
        data_rows = normalized_rows[1:]
        max_cols = max(len(header_source), *(len(row) for row in data_rows), 0)
        header_names = [
            str(header_source[idx]).strip() if idx < len(header_source) and str(header_source[idx]).strip() else f"欄位{idx + 1}"
            for idx in range(max_cols)
        ]
    else:
        max_cols = max(len(row) for row in normalized_rows)
        header_names = [f"欄位{idx + 1}" for idx in range(max_cols)]

    preview_rows: list[dict] = []
    for row in data_rows[:max_preview_rows]:
        row_obj = {}
        for idx, name in enumerate(header_names):
            value = row[idx] if idx < len(row) else None
            row_obj[name] = value
        preview_rows.append(row_obj)

    columns: list[dict] = []
    for idx, name in enumerate(header_names):
        sample_values = []
        for row in data_rows[:8]:
            sample_values.append(row[idx] if idx < len(row) else None)
        columns.append({
            "index": idx,
            "name": name,
            "sample_values": sample_values,
        })

    row_count = len(data_rows)
    return columns, preview_rows, row_count


# ============================================================================
# Endpoints
# ============================================================================

@router.post("/import/preview", response_model=ImportPreviewResponse)
async def import_preview(
    file: UploadFile = File(...),
    data_source: str = Query(""),
    has_header: bool = Query(True)
):
    """
    Preview import file to show preview rows and statistics.
    
    Supports: .xlsx, .xls, .csv, .txt, .pdf
    """
    # Validate file type
    filename = file.filename or "upload"
    file_ext = filename.split('.')[-1].lower()
    if file_ext not in ['xlsx', 'xls', 'csv', 'txt', 'pdf']:
        raise HTTPException(status_code=400, detail=f"Unsupported file format: {file_ext}")
    
    # Read file
    try:
        content = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {str(e)}")

    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    if len(content) > MAX_IMPORT_FILE_SIZE_BYTES:
        raise HTTPException(
            status_code=413,
            detail=(
                f"File '{filename}' exceeds max size {bytes_to_mb_text(MAX_IMPORT_FILE_SIZE_BYTES)} "
                f"(received {bytes_to_mb_text(len(content))})"
            ),
        )
    
    # Extract preview based on file type
    if file_ext in ['xlsx', 'xls']:
        rows, total, empty = extract_excel_preview(content)
    elif file_ext == 'csv':
        rows, total, empty = extract_csv_preview(content)
    elif file_ext == 'pdf':
        rows, total, empty = extract_pdf_preview(content)
    else:
        # TXT file - simple line splitting
        text = content.decode('utf-8')
        rows = [line.split('\t') for line in text.split('\n') if line.strip()]
        total = len(rows)
        empty = 0
    
    # Build preview response
    columns, preview_rows, row_count = build_table_preview(rows, has_header=has_header, max_preview_rows=5)
    if row_count == 0:
        # fallback to parsed total when there is no recognizable tabular row
        row_count = max(total - 1, 0) if has_header else total

    return ImportPreviewResponse(columns=columns, row_count=row_count, preview_rows=preview_rows)


@router.post("/import/submit")
async def import_submit(
    request: ImportSubmitRequest,
    http_request: Request,
    supabase_client = Depends(get_supabase_client)
):
    """
    Submit import batch for processing.
    
    This endpoint:
    1. Creates import batch in PostgreSQL
    2. Queues documents for ElasticSearch indexing
    3. Returns batch ID for status tracking
    """
    try:
        import_batch_label = request.importBatchLabel or request.batch_label or request.file_name or f"batch-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}"
        data_source = request.dataSource or request.data_source or "people-db-import"
        if request.fieldMapping:
            field_mapping = request.fieldMapping.dict()
        elif request.column_mapping:
            field_mapping = request.column_mapping
        else:
            raise HTTPException(status_code=400, detail="fieldMapping/column_mapping is required")

        imported_by_user_id = http_request.headers.get("x-user-id") or http_request.headers.get("X-User-ID")
        if not imported_by_user_id:
            raise HTTPException(status_code=401, detail="Missing user context")
        
        try:
            # Create import batch in PostgreSQL
            batch_id = await supabase_client.create_import_batch(
                label=import_batch_label,
                data_source=data_source,
                total_records=request.total_rows or 0,
                imported_by_user_id=imported_by_user_id,
                field_mapping=field_mapping
            )

            logger.info(f"Created import batch {batch_id} with label '{import_batch_label}'")
            return {
                "batch_id": batch_id,
                "batchId": batch_id,
                "status": "processing",
                "message": f"Import batch queued for processing. Batch ID: {batch_id}",
            }
        except Exception as db_error:
            logger.exception("Import batch creation failed: {}", db_error)
            raise HTTPException(
                status_code=503,
                detail="Import batch tracking is unavailable. Please verify database schema and retry.",
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Import submit failed: {}", e)
        raise HTTPException(status_code=500, detail="Import submit failed")


@router.get("/import/status/{batch_id}")
async def import_status(
    batch_id: str,
    db_client = Depends(get_people_db_client),
    supabase_client = Depends(get_supabase_client)
):
    """Get status of import batch"""
    try:
        # Query import batch status from PostgreSQL
        status_data = await supabase_client.get_import_status(batch_id)
        
        return {
            "batchId": status_data['batch_id'],
            "status": status_data['status'],
            "processed": status_data['processed_records'],
            "total": status_data['total_records'],
            "percentage": status_data['percentage'],
            "errorCount": status_data['error_records'],
            "errorMessage": status_data['error_message'],
            "startedAt": status_data['started_at'],
            "completedAt": status_data['completed_at']
        }
    except Exception as e:
        logger.error(f"Status check failed: {e}", exc_info=True)
        if "not found" in str(e).lower():
            raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/search", response_model=PeopleSearchResponse)
async def search_people(
    q: Optional[str] = Query(None, description="Search query", max_length=256),
    data_source: Optional[str] = Query(None, description="Single data source filter"),
    data_sources: Optional[List[str]] = Query(None, description="Filter by data source"),
    quality: Optional[str] = Query(None, description="Quality band: high/medium/low"),
    imported_from: Optional[str] = Query(None, description="Filter by imported date from"),
    imported_to: Optional[str] = Query(None, description="Filter by imported date to"),
    min_quality_score: Optional[float] = Query(None, ge=0, le=1),
    min_ocr_confidence: float = Query(0.0, ge=0, le=1),
    exclude_duplicates: bool = Query(True),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    limit: Optional[int] = Query(None, ge=1, le=100),
    offset: Optional[int] = Query(None, ge=0),
    db_client = Depends(get_people_db_client)
):
    """
    Search people database with filters.
    
    Supports fuzzy matching across name, ID number, phone, and address fields.
    """
    try:
        effective_limit = limit or page_size
        effective_offset = offset if offset is not None else (page - 1) * effective_limit
        effective_data_sources = [source for source in (data_sources or []) if source]
        if data_source:
            effective_data_sources.append(data_source)
        # deduplicate while preserving order
        effective_data_sources = list(dict.fromkeys(effective_data_sources))

        quality_min, quality_max = resolve_quality_thresholds(quality)
        min_quality = min_quality_score if min_quality_score is not None else quality_min

        filters = {
            'data_sources': effective_data_sources or None,
            'imported_from_date': imported_from,
            'imported_to_date': imported_to,
            'min_quality_score': min_quality,
            'max_quality_score': quality_max,
            'min_ocr_confidence': min_ocr_confidence,
            'exclude_duplicates': exclude_duplicates
        }
        
        result = await db_client.search_documents(
            query_text=q,
            filters=filters,
            limit=effective_limit,
            offset=effective_offset
        )
        
        # Parse results
        hits = result.get('hits', {}).get('hits', [])
        search_results = []
        
        for hit in hits:
            source = hit.get('_source', {})
            full_name = source.get('full_name') or source.get('name') or ''
            search_results.append(PeopleSearchResult(
                record_id=source.get('record_id'),
                full_name=full_name,
                id_number=source.get('id_number'),
                phone=source.get('phone'),
                mobile=source.get('mobile'),
                email=source.get('email'),
                address=source.get('address'),
                company=source.get('company') or source.get('organization'),
                data_source=source.get('data_source') or 'unknown',
                quality_score=source.get('quality_score', 0.0),
                ocr_confidence=source.get('ocr_confidence', 1.0),
                similarity=hit.get('_score', 0.0),
                import_batch_id=source.get('import_batch_id'),
                source_file_path=source.get('source_file_path'),
                source_document_id=source.get('source_document_id'),
                created_at=source.get('created_at'),
            ))
        
        return PeopleSearchResponse(
            total=result.get('hits', {}).get('total', {}).get('value', 0),
            results=search_results,
            page=(effective_offset // effective_limit) + 1 if effective_limit else 1,
            page_size=effective_limit,
            took_ms=result.get('took', 0)
        )
    except Exception as e:
        logger.error(f"Search failed: {e}")
        raise HTTPException(status_code=500, detail="Search failed")


@router.get("/datasets", response_model=DataSourceFacetsResponse)
async def get_data_source_facets(db_client = Depends(get_people_db_client)):
    """Get indexed datasets with document counts for search scope filters."""
    try:
        facets = await db_client.get_data_source_facets()
        return DataSourceFacetsResponse(
            datasets=[
                DataSourceFacet(key=item["key"], count=item["count"])
                for item in facets
            ]
        )
    except Exception as e:
        logger.error(f"Failed to fetch data source facets: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch data source facets")


@router.get("/import/batches")
async def list_import_batches(
    limit: int = Query(20, ge=1, le=200),
    supabase_client = Depends(get_supabase_client)
):
    """List latest import batches for import history visualization."""
    try:
        rows = await supabase_client.list_recent_import_batches(limit=limit)
        return {
            "batches": [
                ImportBatchSummary(
                    batch_id=row.get('id'),
                    label=row.get('label'),
                    data_source=row.get('data_source'),
                    status=row.get('status'),
                    total_records=row.get('total_records') or 0,
                    processed_records=row.get('processed_records') or 0,
                    skipped_records=row.get('skipped_records') or 0,
                    imported_by=row.get('imported_by'),
                    created_at=row.get('created_at'),
                    updated_at=row.get('updated_at'),
                    error_message=row.get('error_message'),
                ).dict()
                for row in rows
            ]
        }
    except Exception as e:
        logger.error(f"Failed to list import batches: {e}")
        raise HTTPException(status_code=500, detail="Failed to list import batches")


@router.get("/stats")
async def get_stats(db_client = Depends(get_people_db_client)):
    """Get index statistics"""
    try:
        stats = await db_client.get_stats()
        return stats
    except Exception as e:
        logger.error(f"Failed to get stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to get stats")
