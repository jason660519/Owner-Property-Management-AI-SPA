#!/usr/bin/env python3
"""
Analyze Excel file to extract table definitions from "各類資料表+RBAC" sheet
"""
import sys
import json

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Please install it first.")
    print("Run: pip3 install --break-system-packages openpyxl")
    sys.exit(1)

def analyze_excel(excel_path, sheet_name="各類資料表+RBAC"):
    """Extract all table definitions from the Excel sheet"""
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Print available sheets
        print("=== Available Sheets ===")
        for sheet in wb.sheetnames:
            print(f"  - {sheet}")
        print()
        
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found!")
            return
        
        ws = wb[sheet_name]
        
        # Extract table information
        print(f"=== Analyzing '{sheet_name}' ===")
        print(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}")
        print()
        
        # Read first few rows to understand structure
        print("=== First 10 rows with column headers ===")
        headers = []
        for col_idx in range(1, min(15, ws.max_column + 1)):
            header_val = ws.cell(row=1, column=col_idx).value
            headers.append(str(header_val) if header_val else f"Col{col_idx}")
        
        print(f"Headers: {' | '.join(headers)}")
        print()
        
        for row_idx in range(2, min(12, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(15, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    row_data.append(str(cell_value)[:30])  # Truncate long values
                else:
                    row_data.append("")
            print(f"Row {row_idx}: {' | '.join(row_data)}")
        
        # Try to extract all table names
        # Need to understand the correct column mapping
        print("\n=== Extracting Table Names ===")
        tables = []
        
        # Based on header analysis
        for row_idx in range(2, ws.max_row + 1):  # Skip header
            # Read all relevant columns
            col1 = ws.cell(row=row_idx, column=1).value  # 工程編號
            col2 = ws.cell(row=row_idx, column=2).value  # 急迫性
            col3 = ws.cell(row=row_idx, column=3).value  # 功能需求編號
            col4 = ws.cell(row=row_idx, column=4).value  # 分類1
            col5 = ws.cell(row=row_idx, column=5).value  # 分類2
            col8 = ws.cell(row=row_idx, column=8).value  # 表名稱-中文
            col9 = ws.cell(row=row_idx, column=9).value  # 功能名稱-英文
            
            # Check if col3 contains category info
            category = str(col3) if col3 else ""
            
            if col8 or col9:
                tables.append({
                    'row': row_idx,
                    'eng_no': col1,
                    'priority': col2 if col2 else "",
                    'category': category,  # This seems to be the actual category
                    'id_field': col4,  # This is the ID field name
                    'subcategory': col5,
                    'name_cn': col8 if col8 else "",
                    'name_en': col9 if col9 else ""
                })
        
        print(f"\nFound {len(tables)} tables:")
        
        # Group by category (col3)
        super_admin_tables = [t for t in tables if 'Super admin' in str(t['category'])]
        landlord_tables = [t for t in tables if 'Landlord' in str(t['category'])]
        tenant_tables = [t for t in tables if 'Tenant' in str(t['category'])]
        finance_tables = [t for t in tables if 'finance' in str(t['category']).lower()]
        other_tables = [t for t in tables if t not in super_admin_tables + landlord_tables + tenant_tables + finance_tables]
        
        print(f"\n=== Super Admin Tables ({len(super_admin_tables)}) ===")
        for i, table in enumerate(super_admin_tables, 1):
            print(f"{i}. [{table['priority']}] {table['name_cn']} ({table['name_en']}) - {table['id_field']}")
        
        print(f"\n=== Landlord Tables ({len(landlord_tables)}) ===")
        for i, table in enumerate(landlord_tables[:50], 1):
            print(f"{i}. [{table['priority']}] {table['name_cn']} ({table['name_en']}) - {table['id_field']}")
        if len(landlord_tables) > 50:
            print(f"... and {len(landlord_tables) - 50} more")
            
        print(f"\n=== Tenant Tables ({len(tenant_tables)}) ===")
        for i, table in enumerate(tenant_tables[:50], 1):
            print(f"{i}. [{table['priority']}] {table['name_cn']} ({table['name_en']}) - {table['id_field']}")
        if len(tenant_tables) > 50:
            print(f"... and {len(tenant_tables) - 50} more")
            
        print(f"\n=== Finance Tables ({len(finance_tables)}) ===")
        for i, table in enumerate(finance_tables[:30], 1):
            print(f"{i}. [{table['priority']}] {table['name_cn']} ({table['name_en']}) - {table['id_field']}")
        if len(finance_tables) > 30:
            print(f"... and {len(finance_tables) - 30} more")
            
        print(f"\n=== Other Tables ({len(other_tables)}) ===")
        for i, table in enumerate(other_tables[:30], 1):
            print(f"{i}. [{table['priority']}] {table['category']}: {table['name_cn']} ({table['name_en']})")
        if len(other_tables) > 30:
            print(f"... and {len(other_tables) - 30} more")
        
        if len(tables) > 50:
            print(f"... and {len(tables) - 50} more tables")
        
        # Save to JSON
        output_file = 'excel_tables_analysis.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump({
                'total_tables': len(tables),
                'super_admin_count': len(super_admin_tables),
                'landlord_count': len(landlord_tables),
                'tenant_count': len(tenant_tables),
                'finance_count': len(finance_tables),
                'other_count': len(other_tables),
                'super_admin_tables': super_admin_tables,
                'landlord_tables': landlord_tables,
                'tenant_tables': tenant_tables,
                'finance_tables': finance_tables,
                'other_tables': other_tables,
                'sheet_name': sheet_name,
                'max_row': ws.max_row,
                'max_column': ws.max_column
            }, f, ensure_ascii=False, indent=2)
        
        print(f"\n✓ Analysis saved to {output_file}")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    excel_path = "Owner Property Management AI Project.xlsx"
    analyze_excel(excel_path)
