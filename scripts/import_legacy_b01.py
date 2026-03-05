#!/usr/bin/env python3
"""
Import legacy B01委託物件資料.xlsx into Supabase local DB.
Tables: property_sales, property_rentals, property_environment_conditions,
        property_owners, property_agent_assignments
"""
import datetime
import uuid
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

DB_URL = "postgresql://postgres:postgres@127.0.0.1:54322/postgres"
EXCEL_PATH = "resources/samples/特力屋EXCEL檔匯出備份/B01委託物件資料.xlsx"
LEGACY_OWNER_ID = "2cd70d9d-9d84-4d2a-9848-df5b3898e4c4"

STATUS_MAP = {
    "逾期": "expired",
    "成交": "sold",
    "委託中": "for_sale",
    "有效": "for_sale",
    "終止": "invalid",
    "暫緩": "pending",
    "保留": "pending",
}


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


def parse_bool(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    s = str(val).strip()
    return s in ("1", "是", "Y", "y", "true", "True")


def parse_num(val, default=0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def parse_int(val, default=0):
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def safe_str(val):
    if val is None:
        return None
    return str(val).strip() or None


def map_gender(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ("男", "M", "male"):
        return "male"
    if s in ("女", "F", "female"):
        return "female"
    return "other"


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    # ---------- Sheet 1: 委託基本資料 ----------
    ws1 = wb["委託基本資料"]
    rows1 = list(ws1.iter_rows(min_row=1, values_only=True))
    headers1 = rows1[0]
    col1 = {h: i for i, h in enumerate(headers1) if h}
    data1 = rows1[1:]

    # ---------- Sheet 2: 環境條件 ----------
    ws2 = wb["環境條件"]
    rows2 = list(ws2.iter_rows(min_row=1, values_only=True))
    headers2 = rows2[0]
    col2 = {h: i for i, h in enumerate(headers2) if h}
    env_map = {}
    for row in rows2[1:]:
        code = row[col2["物件代號"]]
        if code is not None:
            env_map[code] = row

    # ---------- Sheet 3: 屋主資料 ----------
    ws3 = wb["屋主資料"]
    rows3 = list(ws3.iter_rows(min_row=1, values_only=True))
    headers3 = rows3[0]
    col3 = {h: i for i, h in enumerate(headers3) if h}
    owner_map = {}
    for row in rows3[1:]:
        code = row[col3["物件代號"]]
        if code is not None:
            owner_map[code] = row

    wb.close()

    # Clear old imported data
    cur.execute("DELETE FROM property_agent_assignments")
    cur.execute("DELETE FROM property_owners")
    cur.execute("DELETE FROM property_environment_conditions")
    cur.execute("DELETE FROM property_sales WHERE legacy_code IS NOT NULL")
    cur.execute("DELETE FROM property_rentals WHERE legacy_code IS NOT NULL")
    conn.commit()

    sale_count = 0
    rental_count = 0
    env_count = 0
    owner_count = 0
    agent_count = 0

    for row in data1:
        g = lambda key, default=None: row[col1[key]] if key in col1 else default

        legacy_code = g("物件代號")
        if legacy_code is None:
            continue

        data_type = safe_str(g("資料類別"))
        is_sale = data_type != "出租"

        raw_status = safe_str(g("處理狀況")) or ""
        status = STATUS_MAP.get(raw_status, "expired")
        if is_sale and status == "for_sale" and raw_status == "有效":
            status = "for_sale"
        if not is_sale:
            if status == "for_sale":
                status = "for_rent"

        city = safe_str(g("物件地址(縣市)"))
        district = safe_str(g("物件地址(鄉鎮市區)"))
        street = safe_str(g("街路.段.巷.弄"))
        number = safe_str(g("號.樓.室"))
        address = " ".join(filter(None, [city, district, street, number]))
        title = safe_str(g("物件名稱")) or address

        prop_id = str(uuid.uuid4())

        base_cols = {
            "id": prop_id,
            "owner_id": LEGACY_OWNER_ID,
            "address": address or "未知地址",
            "status": status,
            "details": "{}",
            "title": title,
            "address_city": city,
            "address_district": district,
            "address_street": street,
            "address_number": number,
            "legacy_code": safe_str(legacy_code),
            "legacy_key_code": safe_str(g("KEY編號")),
            "contract_category": safe_str(g("契約類別")),
            "custom_number": safe_str(g("自訂編號")),
            "original_listing_code": safe_str(g("原募集物件代號")),
            "branch_code": safe_str(g("分店代號")),
            "building_type": safe_str(g("型式")),
            "building_purpose": safe_str(g("房屋用途")),
            "building_age_years": parse_num(g("屋齡(年)"), None),
            "completion_date_raw": safe_str(g("竣工日期")),
            "orientation": safe_str(g("座向(朝)")),
            "is_corner_unit": parse_bool(g("是否邊間")),
            "construction_company": safe_str(g("建設公司")),
            "building_condition": safe_str(g("屋況")),
            "area_registered": parse_num(g("登記坪數")),
            "area_usable": parse_num(g("使用坪數")),
            "area_main_building": parse_num(g("主建物坪數")),
            "area_auxiliary": parse_num(g("附屬建物坪數")),
            "area_land": parse_num(g("土地坪數")),
            "area_common": parse_num(g("公設坪數")),
            "area_extension": parse_num(g("加蓋坪數")),
            "area_basement": parse_num(g("地下室坪數")),
            "area_shared": parse_num(g("持分坪數")),
            "area_building": parse_num(g("建物坪數")),
            "area_parking": parse_num(g("車位坪數")),
            "basement_registration_note": safe_str(g("地下室登記(使用)")),
            "floor_min": parse_int(g("樓層(1)"), None),
            "floor_max": parse_int(g("樓層(2)(~ 樓)"), None),
            "total_floors": parse_int(g("樓高"), None),
            "basement_floors": parse_int(g("地下層")),
            "layout_rooms": parse_int(g("格局(房)")),
            "layout_living_rooms": parse_int(g("格局(廳)")),
            "layout_bathrooms": parse_int(g("格局(衛)")),
            "layout_balconies": parse_int(g("格局(陽台)")),
            "down_payment": parse_num(g("自備款(萬)")),
            "deposit_amount": parse_num(g("押租(萬)")),
            "current_mortgage_amount": parse_num(g("現有設定額度")),
            "current_loan_amount": parse_num(g("現有貸款額度")),
            "unit_price_registered": parse_num(g("權狀單價(萬/坪)"), None),
            "unit_price_usable": parse_num(g("使用單價(萬/坪)"), None),
            "agent_price": parse_num(g("業價"), None),
            "parking_type": safe_str(g("車位型式")),
            "parking_number": safe_str(g("車位編號")),
            "has_parking": parse_bool(g("是否有車位")),
            "price_includes_parking": parse_bool(g("坪數是否含車位")),
            "parking_price": parse_num(g("車位價")),
            "parking_ownership": safe_str(g("車位產權")),
            "parking_inclusion_note": safe_str(g("含車位情況")),
            "postal_code": safe_str(g("郵遞區號")),
            "latitude": parse_num(g("地圖座標(緯度)"), None),
            "longitude": parse_num(g("地圖座標(經度)"), None),
            "nearby_junior_high": safe_str(g("鄰近國中")),
            "nearby_elementary": safe_str(g("鄰近國小")),
            "nearby_mrt": safe_str(g("捷運站牌")),
            "nearby_park": safe_str(g("公園綠地")),
            "contract_expiry_date": parse_date(g("簽約到期日")),
            "commission_date": parse_date(g("委託日期")),
            "filing_date": parse_date(g("建檔日")),
            "open_date": parse_date(g("開放日期")),
            "last_followup_date": parse_date(g("最後維繫日")),
            "last_registration_date": parse_date(g("上次登記日期")),
            "has_video": parse_bool(g("是否上傳影音檔")),
            "has_floor_plan": parse_bool(g("是否上傳格局圖")),
            "has_property_photos": parse_bool(g("是否上傳房屋照片")),
            "is_onsite_exposure": parse_bool(g("是否現場曝光")),
            "youtube_code": safe_str(g("youtube碼")),
            "vr_720_url": safe_str(g("720VR連結網址")),
            "is_data_open": parse_bool(g("資料是否開放")),
            "main_judgment": safe_str(g("主判")),
            "workflow_control": safe_str(g("作業流程控管")),
            "rent_includes_tax": parse_bool(g("租金是否含稅")) if g("租金是否含稅") else None,
            "rental_deposit_months": parse_int(g("租屋押金(月)"), None),
            "business_area_code": safe_str(g("商圈代號")),
            "business_area_name": safe_str(g("商圈名稱")),
            "community_code": safe_str(g("社區代號")),
            "community_name": safe_str(g("社區名稱")),
            "notes": safe_str(g("備註")),
            "confidential_notes": safe_str(g("機密備註")),
            "property_source": safe_str(g("物件來源")),
            "data_source": safe_str(g("資料來源")),
            "selling_reason": safe_str(g("售屋原因")),
            "advantages": safe_str(g("優點")),
            "disadvantages": safe_str(g("缺點")),
            "is_investigation_signed": parse_bool(g("是否回簽產調書")),
            "transcript_case_number": safe_str(g("謄本轉入原始案號")),
            "last_modifier": safe_str(g("最後修改者")),
            "case_count": parse_int(g("案件數"), 1),
            "extra_rooms": parse_int(g("+房")),
        }

        if is_sale:
            base_cols["price"] = parse_num(g("售價(萬)"))
            cols = list(base_cols.keys())
            vals = [base_cols[c] for c in cols]
            placeholders = ", ".join(["%s"] * len(cols))
            sql = f"INSERT INTO property_sales ({', '.join(cols)}) VALUES ({placeholders})"
            cur.execute(sql, vals)
            sale_count += 1
            prop_type = "sale"
        else:
            base_cols["monthly_rent"] = parse_num(g("押租(萬)"))
            cols = list(base_cols.keys())
            vals = [base_cols[c] for c in cols]
            placeholders = ", ".join(["%s"] * len(cols))
            sql = f"INSERT INTO property_rentals ({', '.join(cols)}) VALUES ({placeholders})"
            cur.execute(sql, vals)
            rental_count += 1
            prop_type = "rental"

        # --- Environment conditions ---
        env_row = env_map.get(legacy_code)
        if env_row:
            eg = lambda key, default=None: env_row[col2[key]] if key in col2 else default
            env_data = {
                "property_id": prop_id,
                "property_type": prop_type,
                "road_width_meters": parse_num(eg("路寬(米)"), None),
                "showing_method": safe_str(eg("帶看方式")),
                "bus_stop": safe_str(eg("公車站牌")),
                "nearby_market": safe_str(eg("鄰近市場")),
                "living_circle": safe_str(eg("生活圈")),
                "highlight_notes": safe_str(eg("重點補充")),
                "ad_copy_40char": safe_str(eg("廣告重點(限40字)")),
                "building_structure": safe_str(eg("建物結構")),
                "building_exterior": safe_str(eg("建物外觀")),
                "building_interior_wall": safe_str(eg("建物內牆")),
                "ceiling_material": safe_str(eg("天花板")),
                "floor_material": safe_str(eg("地板")),
                "lighting_direction": safe_str(eg("採光面")),
                "elevator_households": parse_int(eg("電梯戶數")),
                "elevator_count": parse_int(eg("電梯數量")),
                "management_fee_unit": safe_str(eg("管理費單位")),
                "resident_management_fee": parse_num(eg("住戶管理費(元)")),
                "parking_management_fee": parse_num(eg("車位管理費(元)")),
                "management_fee_includes_parking": parse_bool(eg("管理費是否含車位")),
                "has_security_guard": parse_bool(eg("警衛")),
                "is_currently_leased": parse_bool(eg("是否租賃")),
                "handover_condition": safe_str(eg("交屋情形")),
                "has_natural_gas": parse_bool(eg("是否有天然瓦斯")),
                "area_first_floor": parse_num(eg("一樓面積")),
                "area_second_floor": parse_num(eg("二樓面積")),
                "area_third_floor": parse_num(eg("三樓面積")),
                "area_fourth_floor": parse_num(eg("四樓面積")),
                "area_fifth_floor": parse_num(eg("五樓面積")),
                "area_other_floors": parse_num(eg("其餘樓層面積")),
                "area_arcade": parse_num(eg("騎樓面積")),
                "area_indoor": parse_num(eg("室內面積")),
                "frontage_meters": parse_num(eg("面寬"), None),
                "depth_meters": parse_num(eg("深度"), None),
                "land_use_zone": safe_str(eg("土地使用分區")),
                "extension_location": safe_str(eg("增建位置")),
                "has_ground_vegetation": parse_bool(eg("地上種植物")),
                "has_ground_building": parse_bool(eg("地上是否有建物")),
                "key_access_method": safe_str(eg("鑰匙取得方式")),
                "showing_notes": safe_str(eg("帶看注意事項")),
                "has_mortgage_setting": parse_bool(eg("有無設定")),
                "mortgage_bank": safe_str(eg("設定銀行")),
                "announced_land_value": parse_num(eg("公告現值"), None),
                "parking_entrance": safe_str(eg("車位入口")),
                "parking_status": safe_str(eg("車位狀況")),
                "is_alley_rush": parse_bool(eg("是否路巷沖")),
                "is_safe_alley": parse_bool(eg("是否為安全巷")),
            }
            ecols = list(env_data.keys())
            evals = [env_data[c] for c in ecols]
            cur.execute(
                f"INSERT INTO property_environment_conditions ({', '.join(ecols)}) VALUES ({', '.join(['%s']*len(ecols))})",
                evals,
            )
            env_count += 1

        # --- Owner data ---
        owner_row = owner_map.get(legacy_code)
        if owner_row:
            og = lambda key, default=None: owner_row[col3[key]] if key in col3 else default
            owner_name = safe_str(og("屋主姓名"))
            if owner_name:
                odata = {
                    "property_id": prop_id,
                    "property_type": prop_type,
                    "owner_name": owner_name,
                    "owner_gender": map_gender(og("賣主性別")),
                    "owner_birthday": parse_date(og("屋主生日")),
                    "owner_id_number_enc": safe_str(og("屋主身份證號")),
                    "owner_phone_home": safe_str(og("屋主住家電話")),
                    "owner_phone_office": safe_str(og("屋主公司電話")),
                    "owner_mobile": safe_str(og("屋主手機")),
                    "owner_email": safe_str(og("屋主Email")),
                    "owner_contact_postal_code": safe_str(og("屋主聯絡地址(郵遞區號)")),
                    "owner_residence_postal_code": safe_str(og("屋主戶籍郵遞區號")),
                    "owner_contact_address": safe_str(og("屋主聯絡地址")),
                    "owner_residence_address": safe_str(og("屋主戶籍地址")),
                    "proxy_name": safe_str(og("代理人姓名")),
                    "proxy_phone_home": safe_str(og("代理人住家電話")),
                    "proxy_phone_office": safe_str(og("代理人公司電話")),
                    "proxy_mobile": safe_str(og("代理人手機")),
                    "proxy_birthday": parse_date(og("代理人生日")),
                    "proxy_id_number_enc": safe_str(og("代理人身份證號")),
                    "is_target_customer": parse_bool(og("是否為目標客戶")),
                    "is_active_customer": parse_bool(og("是否為有效客戶")),
                    "can_receive_email": parse_bool(og("是否具收Email")),
                    "notes": safe_str(og("備註")),
                }
                ocols = list(odata.keys())
                ovals = [odata[c] for c in ocols]
                cur.execute(
                    f"INSERT INTO property_owners ({', '.join(ocols)}) VALUES ({', '.join(['%s']*len(ocols))})",
                    ovals,
                )
                owner_count += 1

        # --- Agent assignments (up to 5 slots) ---
        for slot in range(1, 6):
            agent_code_key = f"營業員代號({slot})" if slot == 1 else f"業務員代號({slot})"
            agent_name_key = f"業務員({slot})"
            dept_key = f"部門組別代號({slot})"
            ratio_key = f"件數比率{slot}(%)"

            a_code = safe_str(g(agent_code_key))
            a_name = safe_str(g(agent_name_key))
            if a_code or a_name:
                adata = {
                    "property_id": prop_id,
                    "property_type": prop_type,
                    "slot_number": slot,
                    "agent_code": a_code,
                    "agent_name": a_name,
                    "department_group_code": safe_str(g(dept_key)),
                    "commission_ratio_pct": parse_num(g(ratio_key)),
                }
                acols = list(adata.keys())
                avals = [adata[c] for c in acols]
                cur.execute(
                    f"INSERT INTO property_agent_assignments ({', '.join(acols)}) VALUES ({', '.join(['%s']*len(acols))})",
                    avals,
                )
                agent_count += 1

    conn.commit()
    cur.close()
    conn.close()

    print(f"✅ Import complete!")
    print(f"   property_sales:                  {sale_count}")
    print(f"   property_rentals:                {rental_count}")
    print(f"   property_environment_conditions: {env_count}")
    print(f"   property_owners:                 {owner_count}")
    print(f"   property_agent_assignments:      {agent_count}")


if __name__ == "__main__":
    main()
