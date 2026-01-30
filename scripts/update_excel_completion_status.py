#!/usr/bin/env python3
"""
更新 Excel 中的表格完成状态
在「各類資料表+RBAC」sheet 中添加完成状态列
"""
import openpyxl
from datetime import datetime
import json

# 已完成的表格（从 migration 文件中提取）
completed_tables = {
    # Super Admin - 原始 migration (Gemini)
    'roles', 'permissions', 'role_permissions',
    'platform_settings', 'llm_configs', 'seo_configs', 'notification_templates',
    'currencies', 'exchange_rates', 'i18n_glossary', 'regions_settings',
    'whitelist_blacklist', 'rate_limit_configs',
    'audit_logs', 'api_call_logs', 'error_logs', 'system_maintenance_logs',
    'backup_restore_logs', 'cloud_resources_monitoring', 'ai_performance_metrics', 'web_analytics',
    
    # Super Admin - 補充 migration (Claude)
    'users_track_history', 'tax_rates', 'webhook_configs', 'elasticsearch_indices',
    'perf_metrics', 'recommendation_logs', 'unit_conversion_logs', 'version_history',
    
    # Common User migration
    'user_sessions', 'messages', 'email_threads', 'notification_queue',
    'notification_preferences', 'document_uploads', 'upload_progress',
    'media_processing_queue', 'theme_settings', 'social_auth_connections',
    'calendar_events', 'todo_tasks', 'draft_autosave', 'user_activity_logs',
    'user_feedback',
    
    # Special Features migration
    'virtual_phone_numbers', 'call_logs', 'ai_conversations', 'contracted_tenants',
    'leads_tenants', 'contracted_buyers', 'leads_buyers', 'tenant_inquiries',
    'buyer_inquiries', 'viewing_appointments_tenant', 'viewing_appointments_buyer',
    'lease_agreements', 'sales_agreements', 'deposit_receipts', 'earnest_money_receipts',
    'digital_signatures', 'service_providers', 'maintenance_vendors', 'maintenance_quotes',
    'escrow_legal_services', 'insurance_plans', 'interior_designers', 'user_favorites',
    'property_comparisons', 'user_reviews', 'vlm_parsing_logs',
    
    # Full schema (original tables)
    'users_profile', 'agents', 'agent_authorizations',
    'properties_for_rent', 'properties_for_sale',
    'property_amenities', 'property_photos', 'property_videos',
    'contracts', 'contract_addendums', 'payments',
    'maintenance_records', 'tenant_applications', 'tenant_screening',
    'property_documents', 'landlord_documents', 'tenant_documents',
    'notifications', 'ai_chat_sessions', 'ai_chat_messages',
    'activity_logs', 'system_settings', 'email_templates'
}

def update_excel_status(excel_path, sheet_name="各類資料表+RBAC"):
    """更新 Excel 中的完成状态"""
    
    print("=== 開始更新 Excel 文件 ===\n")
    
    try:
        # 读取 Excel
        wb = openpyxl.load_workbook(excel_path)
        
        if sheet_name not in wb.sheetnames:
            print(f"❌ 錯誤：找不到 sheet '{sheet_name}'")
            return
        
        ws = wb[sheet_name]
        print(f"✓ 成功打開 sheet: {sheet_name}")
        print(f"  行數: {ws.max_row}, 列數: {ws.max_column}\n")
        
        # 找到或创建"完成状态"列
        # 假设表头在第1行，我们在最后一列后面添加状态列
        status_col = ws.max_column + 1
        migration_col = ws.max_column + 2
        executor_col = ws.max_column + 3
        
        # 添加表头
        ws.cell(row=1, column=status_col, value="完成狀態")
        ws.cell(row=1, column=migration_col, value="Migration 文件")
        ws.cell(row=1, column=executor_col, value="執行者")
        
        # 设置表头样式（粗体）
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws.cell(row=1, column=status_col).font = header_font
        ws.cell(row=1, column=status_col).fill = header_fill
        ws.cell(row=1, column=migration_col).font = header_font
        ws.cell(row=1, column=migration_col).fill = header_fill
        ws.cell(row=1, column=executor_col).font = header_font
        ws.cell(row=1, column=executor_col).fill = header_fill
        
        # 遍历所有行，检查完成状态
        completed_count = 0
        pending_count = 0
        
        # 读取分析结果以获取更多信息
        try:
            with open('excel_tables_analysis.json', 'r', encoding='utf-8') as f:
                analysis_data = json.load(f)
        except:
            analysis_data = None
        
        print("開始標記完成狀態...\n")
        
        for row_idx in range(2, ws.max_row + 1):
            # 读取表格英文名称（假设在第9列）
            table_en = ws.cell(row=row_idx, column=9).value
            # 读取ID字段（假设在第4列）
            table_id = ws.cell(row=row_idx, column=4).value
            # 读取中文名称
            table_cn = ws.cell(row=row_idx, column=8).value
            # 读取分类
            category = ws.cell(row=row_idx, column=3).value
            
            if not table_cn and not table_en:
                continue
            
            # 检查是否已完成
            is_completed = False
            migration_file = ""
            executor = ""
            
            if table_en:
                table_en_lower = str(table_en).lower().replace(' ', '_')
                for completed in completed_tables:
                    if completed in table_en_lower:
                        is_completed = True
                        break
            
            if not is_completed and table_id:
                table_id_clean = str(table_id).replace('_id', '').lower()
                for completed in completed_tables:
                    if completed in table_id_clean:
                        is_completed = True
                        break
            
            # 确定 migration 文件和执行者
            if is_completed:
                if category and 'Super admin' in str(category):
                    if table_en and table_en.lower() in ['roles', 'permissions', 'role_permissions', 
                                                           'platform_settings', 'llm_configs', 'seo_configs',
                                                           'currencies', 'exchange_rates', 'i18n_glossary']:
                        migration_file = "20260130_super_admin_tables.sql"
                        executor = "Gemini"
                    else:
                        migration_file = "20260130_super_admin_missing_tables.sql"
                        executor = "Claude"
                elif category and ('所有人都有' in str(category) or 'user' in str(category).lower()):
                    migration_file = "20260130_common_user_tables.sql"
                    executor = "Claude"
                elif category and 'AI Voice' in str(category):
                    migration_file = "20260130_special_features_tables.sql"
                    executor = "Claude"
                elif category and 'customer' in str(category).lower():
                    migration_file = "20260130_special_features_tables.sql"
                    executor = "Claude"
                elif table_en and table_en.lower() in ['users_profile', 'agents', 'properties_for_rent', 'properties_for_sale']:
                    migration_file = "20260122000000_full_schema.sql"
                    executor = "Original Team"
                else:
                    migration_file = "20260130_special_features_tables.sql"
                    executor = "Claude"
                
                status = "✅ 已完成"
                status_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                completed_count += 1
            else:
                status = "⏳ 待完成"
                status_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                migration_file = ""
                executor = ""
                pending_count += 1
            
            # 写入状态
            status_cell = ws.cell(row=row_idx, column=status_col, value=status)
            status_cell.fill = status_fill
            status_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.cell(row=row_idx, column=migration_col, value=migration_file)
            ws.cell(row=row_idx, column=executor_col, value=executor)
        
        # 调整列宽
        ws.column_dimensions[openpyxl.utils.get_column_letter(status_col)].width = 15
        ws.column_dimensions[openpyxl.utils.get_column_letter(migration_col)].width = 40
        ws.column_dimensions[openpyxl.utils.get_column_letter(executor_col)].width = 15
        
        # 在第一个空白行添加统计信息
        summary_row = ws.max_row + 2
        ws.cell(row=summary_row, column=1, value="統計摘要")
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        
        ws.cell(row=summary_row + 1, column=1, value=f"已完成表格數：")
        ws.cell(row=summary_row + 1, column=2, value=completed_count)
        ws.cell(row=summary_row + 1, column=2).font = Font(color="006100", bold=True)
        
        ws.cell(row=summary_row + 2, column=1, value=f"待完成表格數：")
        ws.cell(row=summary_row + 2, column=2, value=pending_count)
        ws.cell(row=summary_row + 2, column=2).font = Font(color="9C5700", bold=True)
        
        ws.cell(row=summary_row + 3, column=1, value=f"總表格數：")
        ws.cell(row=summary_row + 3, column=2, value=completed_count + pending_count)
        ws.cell(row=summary_row + 3, column=2).font = Font(bold=True)
        
        ws.cell(row=summary_row + 4, column=1, value=f"完成率：")
        completion_rate = (completed_count / (completed_count + pending_count) * 100) if (completed_count + pending_count) > 0 else 0
        ws.cell(row=summary_row + 4, column=2, value=f"{completion_rate:.1f}%")
        ws.cell(row=summary_row + 4, column=2).font = Font(bold=True)
        
        ws.cell(row=summary_row + 5, column=1, value=f"更新時間：")
        ws.cell(row=summary_row + 5, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # 保存文件
        backup_path = excel_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        wb.save(backup_path)
        print(f"✓ 備份已保存: {backup_path}\n")
        
        wb.save(excel_path)
        print(f"✓ Excel 文件已更新: {excel_path}\n")
        
        print("=== 更新完成 ===")
        print(f"✅ 已完成: {completed_count} 個表格")
        print(f"⏳ 待完成: {pending_count} 個表格")
        print(f"📊 完成率: {completion_rate:.1f}%")
        print(f"\n已添加以下欄位：")
        print(f"  - 完成狀態（列 {openpyxl.utils.get_column_letter(status_col)}）")
        print(f"  - Migration 文件（列 {openpyxl.utils.get_column_letter(migration_col)}）")
        print(f"  - 執行者（列 {openpyxl.utils.get_column_letter(executor_col)}）")
        
    except Exception as e:
        print(f"❌ 錯誤: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    excel_path = "Owner Property Management AI Project.xlsx"
    update_excel_status(excel_path)
