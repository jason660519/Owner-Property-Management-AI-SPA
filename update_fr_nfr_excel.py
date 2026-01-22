#!/usr/bin/env python3
"""
Update FR+NFR Excel file with comprehensive requirements data
"""

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip3', 'install', 'openpyxl'])
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

# Define the comprehensive FR+NFR requirements data
requirements_data = [
    # Headers
    ["優先級", "需求編號", "分類1", "功能名稱-中文", "功能名稱-英文", "功能說明", "頁面名稱（中文）", "頁面名稱（英文）", 
     "頁面URL尾字母", "可參考的專案頁面", "建造方式／框架／包", "測試方法", "驗收標準與方法", "依賴", "狀態"],
    
    # Priority 1: Company Landing Pages (NFR)
    [1, "001_001", "WebApp-NFR", "公司產品首頁介紹", "Company Homepage", "展示公司產品價值主張與核心功能", "公司首頁", "Home Page", 
     "home_page", "https://www.turbotenant.com", "Expo React Native Web + Next.js", "Lighthouse 效能測試; QA 跨瀏覽器測試", 
     "首頁載入 < 3秒; SEO 優化達標; 響應式設計完整", "", "Not Started"],
    
    [1, "001_002", "WebApp-NFR", "公司產品特色介紹", "Product Features", "詳細展示產品功能特色與優勢", "公司產品特色介紹頁面", "Features Page", 
     "home_page/feature", "/Volumes/KLEVV-4T-1/Real Estate Management Projects/Turbotenant/產品特色.png", "React Components + Animation", 
     "視覺測試; 內容審查", "功能特色清晰展示; 動畫流暢", "", "Not Started"],
    
    [1, "001_003", "WebApp-NFR", "公司產品費用說明", "Pricing Page", "透明展示產品定價方案", "公司產品費用說明頁面", "Pricing Page", 
     "home_page/pricing", "https://www.turbotenant.com/pricing/", "React Pricing Cards", "價格計算邏輯測試", 
     "價格方案清楚; 計算正確", "", "Not Started"],
    
    [1, "001_004", "WebApp-NFR", "公司產品互動教學", "Interactive Tutorial", "提供系統操作指引與教學內容", "互動教學", "Tutorial", 
     "home_page/tutorial", "", "React Joyride 導覽; Supabase 任務追蹤; Lottie 動畫", "單元測試導覽步驟; QA 新帳號 10 分鐘內完成任務; 分析事件驗證", 
     "所有人可於 10 分鐘內完成教學; 完成度事件記錄於分析平台", "", "Not Started"],
    
    [1, "001_005", "WebApp-NFR", "公司產品Q&A+Need Help", "FAQ & Help Center", "提供常見問題解答", "Q&A頁面", "FAQ Page", 
     "home_page/q_and_a", "https://rental.turbotenant.com/landlords/help_center", "React Accordion + 搜尋; Supabase CMS 儲存 FAQ", 
     "單元測試搜尋結果; QA 校對中英文內容; Heatmap 追蹤使用率", "房東可快速搜尋問題; 中英文內容同步; 點擊率追蹤可用", "", "Not Started"],
    
    # Priority 2: Authentication Module (FR)
    [2, "002_001", "WebApp-FR", "登入", "Sign In", "支援手機+Email、社群帳號(Google,Facebook,Apple)、多種註冊方式", "登入頁面", "Sign In Page", 
     "super_admin/authentication/sign_in", "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/auth-signin.html", 
     "Expo React Native 表單 + Supabase Auth API; Twilio Verify 發送 OTP; Expo AuthSession 串接 Google/Facebook", 
     "Jest 表單驗證測試; Supabase sandbox 整合測試; QA 以真實/測試手機與 email 驗證", 
     "使用者可用任一方式註冊; 取得 OTP 後成功登入; Session 於重新整理後仍有效; 僅完成雙重驗證者 status=verified 並可進入 Dashboard", 
     "", "Not Started"],
    
    [2, "002_002", "WebApp-FR", "密碼重設", "Password Reset", "忘記密碼時可透過 Email 重設", "密碼重設頁面", "Password Reset Page", 
     "super_admin/authentication/reset_password", "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/auth-password.html", 
     "Supabase Auth 內建密碼重設流程", "整合測試重設流程", "使用者可成功重設密碼並登入", "002_001", "Not Started"],
    
    [2, "002_003", "WebApp-FR", "註冊", "Sign Up", "Email（預設帳號）+ 手機雙重驗證，標記合格使用者", "註冊頁面", "Sign Up Page", 
     "super_admin/authentication/sign_up", "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/auth-signup.html", 
     "Supabase Edge Function 紀錄驗證狀態; Expo UI 分步驟引導; profiles 表儲存 verified 標記", 
     "整合測試模擬 email/SMS 驗證完成與失敗; QA 驗證未完成時權限受限", 
     "Email/SMS 驗證完成後標記為已驗證; 未驗證時權限受限，導向補驗證流程", "", "Not Started"],
    
    # Priority 3: Super Admin Module (FR)
    [3, "003_001", "WebApp-FR", "最高權限管理員Access Matrix授權設定", "Access Matrix Management", "最高權限管理員Access Matrix－用來對所有人對各項功能的權限設定的管理工具", 
     "權限設定管理頁面", "Access Matrix Page", "super_admin/access_matrix", "", "RBAC + Supabase RLS Policy; React 權限矩陣 UI", 
     "權限邏輯單元測試; 整合測試驗證權限繼承", "管理員可設定各角色權限; 權限變更即時生效", "", "Not Started"],
    
    [3, "003_002", "WebApp-FR", "最高權限管理員－資料庫管理Supabase", "Database Management - Supabase", "Supabase 資料庫管理介面", 
     "Supabase資料庫管理頁面", "Supabase DB Page", "super_admin/database/supabase", "", "Supabase Studio 嵌入; 自定義查詢介面", 
     "資料庫連線測試; 查詢效能測試", "管理員可執行查詢與備份", "003_001", "Not Started"],
    
    [3, "003_003", "WebApp-FR", "最高權限管理員－資料庫管理Elastic Search", "Database Management - ElasticSearch", "ES 資料庫管理介面", 
     "ES資料庫管理頁面", "ElasticSearch Page", "super_admin/database/elastic_search", "", "Kibana 嵌入; ElasticSearch API", 
     "搜尋效能測試; 索引健康檢查", "管理員可管理索引與監控效能", "003_001", "Not Started"],
    
    [3, "003_004", "WebApp-FR", "最高權限管理員的儀表板", "Super Admin Dashboard", "最高權限管理員儀表板頁面", "儀表板頁面", "Admin Dashboard", 
     "super_admin/dashboard", "", "React Dashboard + Chart.js; Supabase 聚合查詢", "資料準確性測試; 效能負載測試", 
     "即時顯示系統統計; 圖表載入 < 2 秒", "003_001", "Not Started"],
    
    [3, "003_005", "WebApp-FR", "最高權限管理員的雲端空間管理功能", "Cloud Storage Management", "最高權限管理員可設置房東的雲端空間大小，流量速度等", 
     "雲端流量控制頁面", "Cloud Traffic Control", "super_admin/cloud_traffic_control", "", "Supabase Storage Quota 管理; CDN 流量監控", 
     "配額設定測試; 流量限制驗證", "可設定房東儲存空間; 流量超限自動通知", "003_001", "Not Started"],
    
    [3, "003_006", "WebApp-FR", "網站效能監控－AI語音回應", "Performance Monitoring - AI Voice", "AI 語音回應延遲 < 2 秒", 
     "效能監控頁面", "Performance Monitor", "super_admin/monitor", "", "語音服務埋點; APM 蒐集 RTA; 邊緣快取語音模型", 
     "自動化語音指令測試; 監控儀表板追蹤延遲; QA 極端環境測試", "AI 語音回應 P95 < 2 秒; 異常自動告警; 儀表板即時更新", 
     "003_004", "Not Started"],
    
    [3, "003_007", "WebApp-FR", "網站效能監控－頁面載入", "Performance Monitoring - Page Load", "監測頁面載入時間 < 3 秒", 
     "效能監控頁面", "Performance Monitor", "super_admin/monitor", "", "前端嵌入 Real User Monitoring; Lighthouse CI Pipeline; New Relic Dashboard", 
     "Lighthouse 監控; k6 壓測; QA 驗證慢速網路測試", "主要頁面在 P95 < 3 秒; 報表可視化; 異常觸發告警", "003_004", "Not Started"],
    
    [3, "003_008", "WebApp-FR", "網站效能監控－故障修復", "Performance Monitoring - MTTR", "網頁故障修復時間 < 4 小時", 
     "效能監控頁面", "Performance Monitor", "super_admin/monitor", "", "PagerDuty 事件流程; 災難復原演練; 狀態頁同步", 
     "定期 DR 演練報告; 模擬故障測試回覆時間; QA 檢查通報", "重大故障 MTTR < 4 小時; 狀態頁 15 分內更新; DR 報告存檔", 
     "003_004", "Not Started"],
    
    [3, "003_009", "WebApp-FR", "網站行為監控與紀錄", "Audit Logging", "確保資料與操作可稽核追蹤", "行為監控頁面", "Audit Log Page", 
     "super_admin/monitor", "", "Supabase 觸發器產生審計紀錄; Cloud Logging 匯整; RLS 權限", 
     "單元測試審計欄位; 安全測試驗證 tamper-proof; 稽核抽查流程", "所有操作具備時間、操作者、變更紀錄; 稽核可匯出 CSV", 
     "003_004", "Not Started"],
    
    # Priority 4: Landlord Core Features - Contracts
    [4, "004_001", "WebApp-FR", "房東的租賃合約製作功能", "Lease Agreement Creator", "線上生成、預覽、儲存租賃合約並支援電子簽名", 
     "租賃合約製作頁面", "Lease Agreement Page", "landlord/dashboard/lease_agreement", "", 
     "模板引擎 (Handlebars) 生成合約; DocuSign 嵌入簽署; Supabase 儲存版本", 
     "單元測試模板佔位符; QA 寄送簽署流程; 自動化測試審核權限", 
     "房東可生成/預覽/寄送合約; 租客可完成電子簽署; 簽署狀態同步", "002_001", "Not Started"],
    
    [4, "004_002", "WebApp-FR", "房東的買賣合約製作功能", "Sale Agreement Creator", "線上生成、預覽、儲存買賣合約並支援電子簽名", 
     "買賣合約製作頁面", "Sale Agreement Page", "landlord/dashboard/sale_agreement", "", 
     "模板引擎 (Handlebars) 生成合約; DocuSign 嵌入簽署; Supabase 儲存版本", 
     "單元測試模板佔位符; QA 寄送簽署流程; 自動化測試審核權限", 
     "房東可生成/預覽/寄送合約; 買方可完成電子簽署; 簽署狀態同步", "002_001", "Not Started"],
    
    [4, "004_003", "WebApp-FR", "房東的儀表板", "Landlord Dashboard", "房東儀表板總覽頁面", "房東儀表板頁面", "Landlord Dashboard", 
     "landlord/dashboard", "", "React Dashboard + Chart.js; Supabase Real-time 訂閱", 
     "資料即時性測試; 響應式佈局測試", "即時顯示物件/租金/預約狀態; 載入 < 2 秒", "002_001", "Not Started"],
    
    [4, "004_004", "WebApp-FR", "房東的客戶-租客篩選功能", "Tenant Screening", "整合信用與背景查核，協助房東評估", 
     "租客篩選頁面", "Tenant Screening Page", "landlord/customer/screen", "", 
     "Supabase 儲存審核報告; 後端整合信用/背景 API; React Native 資料表呈現評分", 
     "Mock API 單元測試; 整合測試驗證風險標籤; QA 交叉檢查不同申請者報告", 
     "房東可查看信用與背景摘要; 系統標記優良/風險租客; 申請列表自動更新", "004_003", "Not Started"],
    
    [4, "004_005", "WebApp-FR", "房東的物件－新增物件", "Add Property", 
     "房東上傳物件文字檔案(Word/PDF/MD)與照片，VLM自動解析並填入欄位；支援手動修正", 
     "新增物件頁面", "Add Property Page", "landlord/properties/add_property", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/property-add.html", 
     "OpenAI Vision API / Claude Vision; Supabase Storage; JSONB 儲存解析結果; 照片自動 CDN 優化", 
     "VLM 解析準確率測試; 違規內容過濾測試; 手動修正流程測試", 
     "支援主流文字與圖片格式; 解析準確率 > 90%; 違規內容自動過濾; 房東可手動修正錯誤", 
     "004_003", "Not Started"],
    
    [4, "004_006", "WebApp-FR", "房東的物件展示功能－Details模式", "Property Details View", "物件詳細資訊頁面", 
     "物件詳情頁", "Property Details", "landlord/properties/details", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/property-details.html", 
     "React Components + Image Gallery", "UI/UX 測試; 資料完整性驗證", "詳細資訊完整展示; 照片輪播流暢", "004_005", "Not Started"],
    
    [4, "004_007", "WebApp-FR", "房東的物件展示功能－Grid模式", "Property Grid View", "物件網格展示", 
     "物件網格頁", "Property Grid", "landlord/properties/grid", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/property-grid.html", 
     "React Grid Layout + Masonry", "響應式佈局測試; 載入效能測試", "Grid 展示流暢; 支援篩選排序", "004_005", "Not Started"],
    
    [4, "004_008", "WebApp-FR", "房東的物件展示功能－List模式", "Property List View", "物件列表展示", 
     "物件列表頁", "Property List", "landlord/properties/list", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/property-list.html", 
     "React Table + 虛擬滾動", "大量資料渲染測試; 排序篩選測試", "列表載入快速; 支援分頁與虛擬滾動", "004_005", "Not Started"],
    
    [4, "004_009", "WebApp-FR", "房東的預約看房管理功能", "Viewing Appointment Management", "接收潛在客戶預約並即時通知房東", 
     "預約看房管理頁面", "Appointment Management", "landlord/properties/appointment", "", 
     "React Native Calendar 元件; Supabase Edge Function 建立/更新預約; SendGrid Email + Twilio SMS 通知", 
     "單元測試檢查時段衝突; 整合測試驗證通知排程; QA 雙重預約與取消情境", 
     "租客可成功預約; 房東於 1 分鐘內收到通知; 狀態同步於租客與房東端顯示", "004_005", "Not Started"],
    
    [4, "004_010", "WebApp-FR", "房東自定義物件的Ｑ＆Ａ功能", "Custom Property FAQ", "房東可自定義房客常見問答", 
     "物件Q&A管理頁", "Property FAQ Page", "landlord/properties/q_a", 
     "https://rental.turbotenant.com/landlords/help_center", 
     "React Native 表單 + Supabase CRUD; 向量資料庫儲存 FAQ 嵌入", 
     "單元測試 CRUD 權限; QA 驗證 FAQ 生效與 AI 回答引用", 
     "房東可新增/編輯/刪除 FAQ; AI 回覆正確引用內容; 版本歷史可追蹤", "004_005", "Not Started"],
    
    [4, "004_011", "WebApp-FR", "房東的行事曆功能", "Calendar Management", "安排看屋,簽約,修理等", 
     "行事曆頁面", "Calendar Page", "landlord/calendar", "Lahomes>Pages>Calendar Page", 
     "React Big Calendar / FullCalendar; Supabase Event 表", "事件 CRUD 測試; 時區處理測試", 
     "可新增/編輯/刪除事件; 支援提醒通知", "004_003", "Not Started"],
    
    [4, "004_012", "WebApp-FR", "簡訊溝通中心", "Message Center", "所有人都可以透過簡訊溝通中心即時或互溝通，或定時聯繫並於預約時間到時手機提醒", 
     "簡訊溝通中心", "Message Center", "message_center", "", 
     "消息佇列 (Supabase Queue) + Twilio SMS; 範本管理 UI", 
     "單元測試排程與節流; QA 發送至測試門號; 自動化測試檢查提醒頻率", 
     "房東可建立與套用通知模板; 系統自動發送重要提醒並紀錄送達狀態", "002_001", "Not Started"],
    
    [4, "004_013", "WebApp-FR", "維修管理", "Maintenance Management (Tradies)", "建立維修申請、與tradies派工聯絡及追蹤tradies進度", 
     "維修派工頁面", "Tradies Management", "landlord/tradies", "", 
     "Supabase 工單表 + 任務狀態機; React Native Kanban UI; Twilio/Email 通知", 
     "單元測試狀態轉換; 整合測試工程師指派; QA 與租客通知流程", 
     "房東可建立工單並指派; 進度與通知即時更新; 歷史紀錄可查", "004_003", "Not Started"],
    
    [4, "004_014", "WebApp-FR", "財務－ATO租賃報稅", "ATO Tax Report", "產生符合澳洲 ATO 的租賃報稅報表", 
     "報稅報表頁面", "Tax Report Page", "landlord/accountant", "", 
     "Supabase Function 計算稅額; PDFKit 產出報表; Email via SendGrid", 
     "單元測試稅額計算; QA 與會計對帳; 自動化測試檢查必填欄位", 
     "可生成 ATO 格式 PDF; 支援 email 與雲端存檔; 數值與實際資料一致", "004_003", "Not Started"],
    
    [4, "004_015", "WebApp-FR", "財務－收支儀表板", "Financial Dashboard", "租金收入、支出與預算儀表板", 
     "財務儀表板", "Financial Dashboard", "landlord/accountant", "", 
     "Supabase 分析視圖; Victory Charts 可視化; PDFKit/Xlsx 輸出", 
     "單元測試資料聚合; Snapshot 測試圖表; QA 匯出 PDF/Excel", 
     "顯示季度/年度收支; 可下載 PDF/Excel; 數據與帳務模組一致", "004_003", "Not Started"],
    
    [4, "004_016", "WebApp-FR", "財務－租金收款管理", "Rent Collection Management", "管理租金收款、欠租提醒與對帳", 
     "租金收款頁面", "Rent Collection", "landlord/accountant", "", 
     "Ledger 資料表 + 定時排程; React Native 財務儀表; 匯出 CSV/PDF", 
     "單元測試帳務計算; 整合測試自動提醒; QA 對帳單匯出與比對", 
     "房東可查看收款紀錄; 設定自動提醒; 系統產出對帳單", "004_003", "Not Started"],
    
    [4, "004_017", "WebApp-FR", "財務－銀行帳戶管理", "Bank Account Management", "綁定多銀行帳戶並指定撥款帳戶", 
     "銀行帳戶管理頁", "Bank Account Page", "landlord/accountant", "", 
     "Plaid Link 嵌入; Supabase 加密儲存銀行資料; 後端排程對帳", 
     "單元測試資料遮罩; Sandbox 銀行整合測試; QA 多帳戶切換撥款", 
     "可新增/刪除多個帳戶; 撥款帳戶可切換; 敏感資料 AES-256 加密", "004_003", "Not Started"],
    
    [4, "004_018", "WebApp-FR", "金流服務－線上支付", "Online Payment", "支援租金線上支付與自動扣款", 
     "線上支付頁面", "Payment Page", "landlord/accountant", "", 
     "Serverless 金流編排; React Native Web Checkout; Vault 儲存 token", 
     "PCI 擬真測試; Sandbox 交易流程; QA 檢查自動扣款與退款", 
     "租客可完成支付與自動扣款; 房東在 Dashboard 可見入帳紀錄; 手續費計算正確", "004_017", "Not Started"],
    
    [4, "004_019", "WebApp-FR", "金流服務－離線付款", "Offline Payment Registration", "支援線下類 Apple Pay 付款登錄", 
     "離線付款登記頁", "Offline Payment Page", "landlord/accountant", "", 
     "NFC/Token API 整合; Supabase 記錄離線交易; 後端批次同步", 
     "單元測試交易同步; QA 使用測試裝置離線付款; 自動化檢查重複上傳", 
     "離線交易可成功註記並同步; 發票與收據自動寄送; 避免重複扣款", "004_018", "Not Started"],
    
    [4, "004_020", "WebApp-FR", "屋主的行銷部落格網站行為監控", "Blog Traffic Analytics", "可以看到潛在客戶的（ip）登入看了屋主哪個物件行銷的blog", 
     "部落格監控頁面", "Blog Monitor", "landlord/blogs/monitor", "", 
     "Google Analytics / Mixpanel; IP 追蹤; Heatmap", "事件追蹤測試; 隱私合規檢查", 
     "可追蹤訪客瀏覽路徑; 符合隱私法規", "004_031", "Not Started"],
    
    [4, "004_021", "WebApp-FR", "房東的email inbox信箱", "Email Inbox", "房東的email inbox信箱", 
     "郵件收件箱", "Email Inbox", "landlord/email_inbox", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/inbox.html", 
     "IMAP/SMTP 整合; React Email Client", "郵件收發測試; 附件處理測試", 
     "可收發郵件; 支援附件; 郵件分類", "004_003", "Not Started"],
    
    [4, "004_022", "WebApp-FR", "房東的仲介－Details模式", "Agent Details View", "仲介詳細資訊", 
     "仲介詳情頁", "Agent Details", "landlord/agents/details", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/agents-details.html", 
     "React Detail Page", "資料完整性測試", "仲介資訊完整展示", "004_003", "Not Started"],
    
    [4, "004_023", "WebApp-FR", "房東的仲介－Grid模式", "Agent Grid View", "仲介網格展示", 
     "仲介網格頁", "Agent Grid", "landlord/agents/grid", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/agents-grid.html", 
     "React Grid Layout", "佈局測試", "仲介 Grid 展示流暢", "004_003", "Not Started"],
    
    [4, "004_024", "WebApp-FR", "房東的仲介－List模式", "Agent List View", "仲介列表展示", 
     "仲介列表頁", "Agent List", "landlord/agents/list", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/agents-list.html", 
     "React Table", "列表測試", "仲介列表可排序篩選", "004_003", "Not Started"],
    
    [4, "004_025", "WebApp-FR", "房東的仲介－新增仲介", "Add Agent", "新增仲介資料", 
     "新增仲介頁", "Add Agent", "landlord/agents/add_agent", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/agents-add.html", 
     "React Form + Supabase CRUD", "表單驗證測試", "可成功新增仲介資料", "004_003", "Not Started"],
    
    [4, "004_026", "WebApp-FR", "房東的客戶－Details模式", "Customer Details View", "客戶詳細資訊", 
     "客戶詳情頁", "Customer Details", "landlord/customers/details", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/customers-details.html", 
     "React Detail Page", "資料完整性測試", "客戶資訊完整展示", "004_003", "Not Started"],
    
    [4, "004_027", "WebApp-FR", "房東的客戶－Grid模式", "Customer Grid View", "客戶網格展示", 
     "客戶網格頁", "Customer Grid", "landlord/customers/grid", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/customers-grid.html", 
     "React Grid Layout", "佈局測試", "客戶 Grid 展示流暢", "004_003", "Not Started"],
    
    [4, "004_028", "WebApp-FR", "房東的客戶－List模式", "Customer List View", "客戶列表展示", 
     "客戶列表頁", "Customer List", "landlord/customers/list", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/customers-list.html", 
     "React Table", "列表測試", "客戶列表可排序篩選", "004_003", "Not Started"],
    
    [4, "004_029", "WebApp-FR", "房東的客戶－新增客戶", "Add Customer", "新增客戶資料", 
     "新增客戶頁", "Add Customer", "landlord/customers/add_customer", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/customers-add.html", 
     "React Form + Supabase CRUD", "表單驗證測試", "可成功新增客戶資料", "004_003", "Not Started"],
    
    [4, "004_030", "WebApp-FR", "房東的客戶－成交客戶", "Closed Customers", "已成交簽約買方或租方管理頁面", 
     "成交客戶管理頁", "Closed Customers", "landlord/customers/buyer_and_renter", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/orders.html", 
     "React Table + Supabase Query", "交易狀態測試", "顯示已成交客戶與合約狀態", "004_001, 004_002", "Not Started"],
    
    [4, "004_031", "WebApp-NFR", "房東的部落格創建功能", "Blog Creator", "房東只要上傳照片＋文字說明，按下Create Blog buttom 就能創建物件或個人的部落格展示頁面", 
     "創建部落格頁", "Create Blog", "landlord/create_blog", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/post-create.html", 
     "React Native Web 模板引擎 + Mustache; Supabase metadata; 靜態頁面托管於 Vercel", 
     "Snapshot 測試模板; Cypress 驗證分享連結; QA 測試資料不足提示", 
     "資料齊全時自動生成分享連結; 缺資料時顯示待補欄位清單", "004_005", "Not Started"],
    
    [4, "004_032", "WebApp-NFR", "全球屋主的物件部落格展示頁面", "Global Property Blogs", "全球物業部落格展示", 
     "全球部落格展示頁", "Global Blogs", "blog_grid/global", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/post.html", 
     "React Blog Grid + Supabase Public Query", "SEO 測試; 載入效能測試", 
     "公開可訪問; SEO 優化; 載入快速", "004_031", "Not Started"],
    
    [4, "004_033", "WebApp-NFR", "全部台灣屋主的物件部落格展示頁面", "Taiwan Property Blogs", "台灣物業部落格展示", 
     "台灣部落格展示頁", "Taiwan Blogs", "blog_grid/tw", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/post.html", 
     "React Blog Grid + 地區篩選", "地區篩選測試", "只顯示台灣物件; 支援中文搜尋", "004_031", "Not Started"],
    
    [4, "004_034", "WebApp-NFR", "房東的溝通頁面", "Communication Page", "未檢視／已檢視的聊天記錄，群聊功能，私聊功能", 
     "房東房客溝通頁面", "Messages Page", "landlord/message", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/messages.html", 
     "React Chat UI + Supabase Realtime; WebSocket 連線", "即時訊息測試; 已讀狀態測試", 
     "訊息即時傳遞; 支援群聊私聊; 已讀未讀標記", "002_001", "Not Started"],
    
    # Priority 8: Third-party Services
    [8, "008_001", "WebApp-NFR", "第三方會計人員查帳審計功能", "Third-party Audit", "第三方會計人員查帳審計功能", 
     "第三方審計頁", "Third-party Audit", "third-party/accountants/audit", "", 
     "只讀權限 RLS Policy; 審計報表匯出", "權限測試; 資料遮罩測試", 
     "第三方可查看授權資料; 不可修改; 可匯出審計報表", "003_001", "Not Started"],
    
    [8, "008_002", "WebApp-NFR", "網路安全－隱私審計", "Privacy Audit", "審計隱私或個人資料有無外洩的可能性", 
     "隱私審計頁", "Privacy Audit", "cyber_security/auditor/private", "", 
     "GDPR 合規檢查; 資料加密驗證; 存取日誌分析", "自動化合規掃描; 滲透測試", 
     "通過隱私合規檢查; 無資料外洩風險", "", "Not Started"],
    
    # Priority 700: Non-Functional Requirements
    [700, "700_001", "WebApp-NFR", "資訊安全－HTTPS與加密", "HTTPS & Encryption", "實施 HTTPS 傳輸、敏感資料加密與定期稽核", 
     "無特定頁面", "N/A", "cyber_security/ssl_lab", "", 
     "全站 HTTPS; KMS 管理密鑰; 年度稽核 SOP", "SSL Labs 測試 A 級; 密鑰輪換測試; 稽核報告", 
     "公司稽核人員可取得完整報告; 加密與稽核紀錄完善", "", "Not Started"],
    
    [700, "700_002", "WebApp-NFR", "資訊安全－防止未授權修改", "RBAC Security", "防止未授權修改個資與合約資料", 
     "無特定頁面", "N/A", "cyber_security/rbac", "", 
     "雙層 RBAC, WAF, 資料庫行級權限; 敏感欄位加密", "滲透測試; 自動化安全掃描; 權限 QA 用例", 
     "未授權行為被阻擋並記錄; 敏感資料加密; 安全稽核通過", "", "Not Started"],
    
    [700, "700_003", "WebApp-NFR", "可學習性", "Learnability", "清楚操作指引，新用戶 10 分鐘內上架物件", 
     "無特定頁面", "N/A", "", "", 
     "Onboarding Checklist + 教學影片; 互動引導", "可用性測試觀察新用戶; 分析追蹤完成時間", 
     "80% 新用戶 10 分鐘內完成首次上架; 引導完成率被記錄", "001_004", "Not Started"],
    
    [700, "700_004", "WebApp-NFR", "可擴展性", "Scalability", "支援水平擴展應對用戶成長", 
     "無特定頁面", "N/A", "", "", 
     "Kubernetes 水平擴展; Supabase 連線池; Queue 解耦", "負載測試逐步放量; Auto-scaling 演練; 容量報告", 
     "系統可在需求成長時自動擴展; 無重大性能衰退", "", "Not Started"],
    
    [700, "700_005", "WebApp-NFR", "可維護性", "Maintainability", "採模組化、前後端分離", 
     "無特定頁面", "N/A", "", "", 
     "Monorepo + Turborepo; CI 靜態檢查; 型別界面", "單元/整合測試涵蓋核心模組; 靜態分析; Code review checklist", 
     "模組邊界清晰; CI < 10 分鐘完成; 重大變更有測試覆蓋", "", "Not Started"],
    
    [700, "700_006", "WebApp-NFR", "易用性", "Usability", "友善 UI，房東可透過語音或點擊操作", 
     "無特定頁面", "N/A", "", "", 
     "Design System + 無障礙規範; Crisp Chat 與語音整合", "可用性測試 SUS ≥ 80; 無障礙自動化掃描", 
     "SUS 分數 ≥ 80; 語音/點擊流程皆完成; 無障礙 AA 達成", "", "Not Started"],
    
    [700, "700_007", "WebApp-NFR", "系統可用性", "System Availability", "服務可用性 99.5%，支援故障轉移", 
     "無特定頁面", "N/A", "", "", 
     "Multi-AZ 部署; 自動故障轉移; 備援資料庫", "Chaos 工程測試; 監控 SLA; DR 演練", 
     "年度可用性 ≥99.5%; 故障轉移 < 2 分鐘; DR 報告有效", "", "Not Started"],
    
    [700, "700_008", "WebApp-NFR", "繁TW／英EN 介面切換", "Multilingual Support", "多語言專業詞彙對照表，以直觀的table方式直接套用管理i18next", 
     "多語言管理頁", "i18next Table", "super_admin/multilanguage/i18next_table", "", 
     "i18next 框架; 專業詞彙對照表; React Table 管理介面", "語言切換測試; 翻譯完整性測試", 
     "多作業系統都能完整展示正確的文字; 即時切換語言", "", "Not Started"],
    
    [700, "700_009", "WebApp-NFR", "響應式設計", "Responsive Design", "支援 Web 與 Mobile 提供一致體驗", 
     "無特定頁面", "N/A", "", "", 
     "Expo React Native Web 響應式格線; Tailwind RN; 裝置偵測", 
     "自動化視窗尺寸測試; QA 真機/模擬器; Lighthouse 行動評分", 
     "不同裝置呈現一致; 行動 Lighthouse 分數 ≥ 90; 主要流程可順利完成", "", "Not Started"],
    
    [700, "700_010", "WebApp-NFR", "全球物業展示頁面", "Global Property Listing", "公開有效出租/出售物件，不顯示過期資訊", 
     "全球物業展示頁", "Global Listings", "", "", 
     "Expo Web/Next.js 公開站; CDN 快取; 排程清除過期物件", 
     "自動化測試驗證過期物件不顯示; Load test 全球節點; SEO 檢查", 
     "只顯示有效物件; 監管合規; 全球載入 P95 < 4 秒", "", "Not Started"],
    
    [700, "700_011", "WebApp-NFR", "全球物業客戶評價Review", "Global Reviews", "全球物業展示頁面Review", 
     "評價頁面", "Reviews Page", "", 
     "file:///Volumes/KLEVV-4T-1/Real Estate Management Projects/Lahomes/techzaa.in/lahomes/admin/reviews.html", 
     "React Review Component + Supabase 評價表", "評價真實性驗證; 惡意評價過濾測試", 
     "顯示真實評價; 支援評分排序; 惡意評價可檢舉", "", "Not Started"],
    
    [700, "700_012", "WebApp-NFR", "3D/2D導覽", "3D/2D Virtual Tour", "提供物件 3D/2D 導覽模式", 
     "3D導覽頁", "Virtual Tour", "", "", 
     "WebGL/Krpano 內嵌; 媒體 CDN; 預載縮圖", "前端 Lighthouse 測速; QA 跨瀏覽器/裝置測試; 壓測載入時間", 
     "VIP 房東可在 5 秒內載入導覽; 支援 3D/2D 切換; FPS 穩定 ≥ 30", "", "Not Started"],
    
    [700, "700_013", "WebApp-NFR", "3D/2D變裝", "AI Interior Redesign", "AI 圖生圖多風格改裝建議", 
     "AI變裝頁", "AI Redesign", "", "", 
     "GPU 推理服務 (Stable Diffusion) + 任務佇列; React 動畫展示", 
     "單元測試任務排程; QA 驗證各風格輸出; 人工審查內容安全", 
     "提供至少 3 個風格版本; 處理時間 < 10 分鐘; 可下載高解析圖", "", "Not Started"],
    
    [700, "700_014", "WebApp-NFR", "720度全景看屋", "360° Virtual Tour", "提供 720 度全景瀏覽", 
     "全景看屋頁", "360 Tour", "", "", 
     "360 Viewer (krpano) + VR 模式; CDN 切片串流", "Lighthouse 與 WebPageTest 效能; QA VR 裝置與手機體驗", 
     "全景載入 < 6 秒; 支援手機陀螺儀; 無效媒體會顯示錯誤提示", "", "Not Started"],
    
    [700, "700_015", "WebApp-NFR", "AI 小幫手", "AI Assistant", "AI 助手解釋頁面、合約、報表內容", 
     "AI助手浮動視窗", "AI Assistant", "", "", 
     "LangChain + OpenAI/Azure GPT; Supabase 向量資料庫; 前端浮動助理", 
     "Prompt 回歸測試; QA 法規/報表問答; 安全過濾測試", 
     "用戶可即時提問; AI 回答引用來源; 正確率 ≥ 90% 且可導出", "", "Not Started"],
    
    [700, "700_016", "WebApp-NFR", "AI講房", "AI Property Narration", "自動解說物件周邊資訊", 
     "AI講房功能", "AI Narration", "", "", 
     "Supabase 地區資料 + Google Places; LLM 摘要服務; React Native 語音播放", 
     "Mock API 單元測試; Prompt 回歸測試; QA 多城市口語輸出", 
     "AI 解說包含交通/生活機能; 內容正確率 ≥ 90%; 支援文字與語音播放", "", "Not Started"],
    
    [700, "700_017", "WebApp-NFR", "加值服務－智能門鎖", "Smart Lock Integration", "多元解鎖與遠端管理", 
     "智能設備管理頁", "Smart Devices", "", "", 
     "鎖控服務 (AWS IoT); AES 加密憑證庫; App 遠端控制 UI", 
     "單元測試 OTP 與一次性密碼; QA 遠端開門與低電量提醒; 安全掃描", 
     "支援指紋/密碼/卡片/App/一次性密碼; 異常即時通知; 電量不足提醒", "", "Not Started"],
    
    [700, "700_018", "WebApp-NFR", "加值服務－保險方案", "Insurance Integration", "提供保險方案申購與保單進度管理", 
     "保險服務頁", "Insurance Page", "", "", 
     "Supabase workflow + 保險 API (Sandbox); React Native 流程引導", 
     "單元測試保單狀態變更; 整合測試申購流程; QA 核對保單文件", 
     "房東可瀏覽方案、提交申購、追蹤進度; 通知與文件歸檔完整", "", "Not Started"],
    
    [700, "700_019", "WebApp-NFR", "加值服務－租金保障", "Rent Guarantee", "提供租金代墊、帳單自動化與債權轉移", 
     "租金保障頁", "Rent Guarantee", "", "", 
     "風險評估引擎 + 金融 API; Supabase 合約紀錄; 自動扣款整合", 
     "情境模擬測試理賠; QA 確認帳單自動化; 合規審查", 
     "房東可啟用保障方案; 代墊/扣款流程透明; 報表可追蹤", "", "Not Started"],
    
    [700, "700_020", "WebApp-NFR", "加值服務－攝影機監控", "Camera Monitoring", "整合攝影機即時監控", 
     "監控頁面", "Camera Monitor", "", "", 
     "ONVIF/RTSP 串流服務; WebRTC Viewer; 加密串流", "安全性滲透測試; QA 多裝置即時串流; 負載測試", 
     "即時影像延遲 < 1 秒; 權限控管依角色; 錄影可回放", "", "Not Started"],
    
    [700, "700_021", "WebApp-NFR", "專屬轉接號碼", "Virtual Phone Number", "提供物件專屬轉接號碼與來電來源顯示", 
     "轉接號碼管理頁", "Virtual Number", "", "", 
     "Serverless Voice 路由 (Twilio); Expo UI 控制開關; 通話紀錄寫入 Supabase", 
     "Sandbox 來電測試; Twilio 通話錄音驗證; QA 測試啟停動作", 
     "房東可啟停轉接號碼; 來電來源顯示與紀錄完整; 通話資訊可匯出", "", "Not Started"],
    
    [700, "700_022", "WebApp-NFR", "影片導覽服務", "Video Tour Service", "協助拍攝與託管導覽影片", 
     "影片管理頁", "Video Tour", "", "", 
     "Mux 上傳 + 自動編解碼; React 播放器組件", "單元測試影片狀態; QA 上傳/播放流程; 合規測試 CDN 權限", 
     "房東可上傳或申請拍攝; 影片可在行銷頁播放; 產出分享連結", "", "Not Started"],
]

def update_excel_file():
    """Update the Excel file with comprehensive FR+NFR data"""
    
    excel_file = "FR+NFR Project Management-Owner, Real Estate Agent SaaS.xlsx"
    
    try:
        # Try to load existing workbook
        wb = load_workbook(excel_file)
        print(f"✓ Loaded existing Excel file: {excel_file}")
        
        # Check if "FR+NFR" sheet exists, if not create it
        if "FR+NFR" in wb.sheetnames:
            # Remove existing sheet
            del wb["FR+NFR"]
            print("  Removed existing 'FR+NFR' sheet")
        
        # Create new sheet
        ws = wb.create_sheet("FR+NFR", 0)  # Insert as first sheet
        print("  Created new 'FR+NFR' sheet")
        
    except FileNotFoundError:
        print(f"✗ File not found: {excel_file}")
        print("  Creating new workbook...")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "FR+NFR"
    
    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Define priority colors
    priority_colors = {
        1: "FFD966",  # Yellow - Highest priority
        2: "FFE699",  # Light Yellow
        3: "BDD7EE",  # Light Blue
        4: "C5E0B4",  # Light Green
        8: "F4B084",  # Light Orange
        700: "E2EFDA"  # Pale Green
    }
    
    # Write data to sheet
    for row_idx, row_data in enumerate(requirements_data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Apply header style
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            else:
                # Apply priority-based color
                if col_idx == 1 and isinstance(cell_value, int):  # Priority column
                    color = priority_colors.get(cell_value, "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                # Apply text wrapping for long content
                if col_idx in [4, 5, 6, 11, 12, 13]:  # Description columns
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Set column widths
    column_widths = {
        'A': 8,   # 優先級
        'B': 12,  # 需求編號
        'C': 15,  # 分類1
        'D': 30,  # 功能名稱-中文
        'E': 30,  # 功能名稱-英文
        'F': 50,  # 功能說明
        'G': 20,  # 頁面名稱（中文）
        'H': 20,  # 頁面名稱（英文）
        'I': 35,  # 頁面URL尾字母
        'J': 60,  # 可參考的專案頁面
        'K': 50,  # 建造方式／框架／包
        'L': 50,  # 測試方法
        'M': 60,  # 驗收標準與方法
        'N': 15,  # 依賴
        'O': 12,  # 狀態
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Set row heights
    ws.row_dimensions[1].height = 40  # Header row
    for row_idx in range(2, len(requirements_data) + 1):
        ws.row_dimensions[row_idx].height = 60
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save the workbook
    wb.save(excel_file)
    print(f"\n✓ Successfully updated Excel file: {excel_file}")
    print(f"  Total requirements: {len(requirements_data) - 1}")  # Exclude header
    print(f"  Sheet name: FR+NFR")
    
    # Print summary by priority
    priority_counts = {}
    for row in requirements_data[1:]:  # Skip header
        priority = row[0]
        priority_counts[priority] = priority_counts.get(priority, 0) + 1
    
    print("\n📊 Requirements Summary by Priority:")
    for priority in sorted(priority_counts.keys()):
        print(f"  Priority {priority}: {priority_counts[priority]} items")

if __name__ == "__main__":
    print("=" * 70)
    print("FR+NFR Excel File Update Script")
    print("=" * 70)
    print()
    
    update_excel_file()
    
    print()
    print("=" * 70)
    print("✓ Update completed successfully!")
    print("=" * 70)
